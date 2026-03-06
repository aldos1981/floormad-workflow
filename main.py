from fastapi import FastAPI, HTTPException, Request, Form, UploadFile, File, Body
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
import os
from dotenv import load_dotenv
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import Optional, Dict, Any, List
import uuid
import json

from database import get_db_connection, init_db
from workflow_engine import WorkflowEngine
from google_auth_oauthlib.flow import Flow
from google.oauth2.credentials import Credentials

# Load environment variables
load_dotenv()

app = FastAPI(title="Floormad Automation Manager")

# CORS (Critical for cross-domain or cloud hosting)
from fastapi.middleware.cors import CORSMiddleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Allow all for now
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Models
class ProjectCreate(BaseModel):
    name: str
    description: Optional[str] = None
    google_sheet_id: Optional[str] = ""
    service_account_json: Optional[str] = ""
    smtp_config: Optional[Dict[str, Any]] = None
    wesendit_config: Optional[Dict[str, Any]] = None
    pipedrive_config: Optional[Dict[str, Any]] = None
    cron_expression: Optional[str] = None
    price_list_url: Optional[str] = None
    locality_prompt: Optional[str] = None
    products_config: Optional[List[Dict[str, Any]]] = None
    workflow_json: Optional[Dict[str, Any]] = None

class ProjectUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    status: Optional[str] = None
    google_sheet_id: Optional[str] = None
    service_account_json: Optional[str] = None
    cron_expression: Optional[str] = None
    price_list_url: Optional[str] = None
    locality_prompt: Optional[str] = None
    products_config: Optional[List[Dict[str, Any]]] = None
    workflow_json: Optional[Dict[str, Any]] = None
    smtp_config: Optional[Dict[str, Any]] = None
    wesendit_config: Optional[Dict[str, Any]] = None
    pipedrive_config: Optional[Dict[str, Any]] = None

from fastapi import WebSocket, WebSocketDisconnect

class ConnectionManager:
    def __init__(self):
        self.active_connections: List[WebSocket] = []

    async def connect(self, websocket: WebSocket):
        await websocket.accept()
        self.active_connections.append(websocket)

    def disconnect(self, websocket: WebSocket):
        self.active_connections.remove(websocket)

    async def broadcast(self, message: str):
        for connection in self.active_connections:
            try:
                await connection.send_text(message)
            except:
                pass

manager = ConnectionManager()


class GlobalSettings(BaseModel):
    service_account_json: Optional[str] = None
    google_api_key: Optional[str] = None
    default_sheet_id: Optional[str] = None
    google_client_id: Optional[str] = None
    google_client_secret: Optional[str] = None

import asyncio
import logging
from datetime import datetime, timedelta

logger = logging.getLogger("cron_scheduler")

# Minimum interval between cron runs for the same project (prevents spam)
MIN_CRON_INTERVAL_MINUTES = 4

# Track last cron run time per project
_last_cron_runs: Dict[str, datetime] = {}

async def cron_scheduler_loop():
    """
    Background task that checks active projects for cron triggers
    and runs their workflows automatically.
    """
    logger.info("🕐 Cron Scheduler started")
    
    # Wait 10 seconds for the app to fully initialize
    await asyncio.sleep(10)
    
    while True:
        try:
            from croniter import croniter
            
            conn = get_db_connection()
            # Get all active projects
            projects = conn.execute(
                "SELECT id, name, status, workflow_json FROM projects WHERE status = 'active'"
            ).fetchall()
            
            # Get global settings for API key
            settings = conn.execute("SELECT key, value FROM settings").fetchall()
            settings_dict = {row['key']: row['value'] for row in settings}
            conn.close()
            
            now = datetime.now()
            
            for project in projects:
                try:
                    workflow_json = project['workflow_json']
                    if not workflow_json:
                        continue
                    
                    workflow_data = json.loads(workflow_json)
                    
                    # Find the TRIGGER node and get its cron_expression
                    cron_expr = None
                    trigger_type = None
                    nodes = workflow_data.get('drawflow', {}).get('Home', {}).get('data', {})
                    
                    for node_id, node in nodes.items():
                        if node.get('name') == 'TRIGGER':
                            config = node.get('data', {}).get('config', {})
                            trigger_type = config.get('trigger_type', 'manual')
                            cron_expr = config.get('cron_expression', '')
                            break
                    
                    # Skip if not a cron trigger or no expression
                    if trigger_type != 'cron' or not cron_expr or not cron_expr.strip():
                        continue
                    
                    # Validate cron expression
                    if not croniter.is_valid(cron_expr):
                        logger.warning(f"Invalid cron expression '{cron_expr}' for project {project['name']}")
                        continue
                    
                    # Check if cron matches current minute
                    cron = croniter(cron_expr, now - timedelta(seconds=60))
                    next_run = cron.get_next(datetime)
                    
                    # Check if the next run is within the current minute
                    current_minute = now.replace(second=0, microsecond=0)
                    next_run_minute = next_run.replace(second=0, microsecond=0)
                    
                    if next_run_minute != current_minute:
                        continue
                    
                    # Check minimum interval between runs (anti-spam)
                    project_id = project['id']
                    last_run = _last_cron_runs.get(project_id)
                    if last_run and (now - last_run).total_seconds() < MIN_CRON_INTERVAL_MINUTES * 60:
                        continue  # Too soon since last run
                    
                    # RUN THE WORKFLOW!
                    _last_cron_runs[project_id] = now
                    logger.info(f"🚀 Cron trigger firing for project '{project['name']}' (cron: {cron_expr})")
                    
                    # Run in background to not block the scheduler
                    asyncio.create_task(_run_cron_workflow(project_id, project['name'], settings_dict))
                    
                except Exception as e:
                    logger.error(f"Error checking cron for project {project.get('name', '?')}: {e}")
            
        except Exception as e:
            logger.error(f"Cron scheduler error: {e}")
        
        # Check every 30 seconds
        await asyncio.sleep(30)


async def _run_cron_workflow(project_id: str, project_name: str, settings_dict: dict):
    """Run a workflow triggered by cron schedule."""
    try:
        conn = get_db_connection()
        project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
        conn.close()
        
        if not project:
            return
        
        workflow_data = json.loads(project['workflow_json'])
        
        engine = WorkflowEngine(
            workflow_data,
            context={"project": dict(project)},
            api_key=settings_dict.get('google_api_key')
        )
        
        # No status_callback for cron runs (no UI to update)
        result = await engine.run(None)
        
        # Log the run
        run_id = str(uuid.uuid4())
        status = result.get('status', 'unknown')
        log_count = len(result.get('log', []))
        
        conn = get_db_connection()
        conn.execute(
            "INSERT INTO runs (id, project_id, timestamp, leads_processed, status, log_details) VALUES (?, ?, ?, ?, ?, ?)",
            (run_id, project_id, datetime.now().isoformat(), log_count, status, json.dumps(result.get('log', []), default=str))
        )
        conn.commit()
        conn.close()
        
        logger.info(f"✅ Cron workflow completed for '{project_name}': {status} ({log_count} nodes)")
        
        # Broadcast update via WebSocket 
        await manager.broadcast(json.dumps({
            "type": "cron_execution_complete",
            "project_id": project_id,
            "project_name": project_name,
            "status": status,
            "nodes_processed": log_count,
            "timestamp": datetime.now().isoformat()
        }))
        
    except Exception as e:
        logger.error(f"❌ Cron workflow failed for '{project_name}': {e}")


@app.on_event("startup")
async def startup_event():
    init_db()
    # Start the cron scheduler as a background task
    asyncio.create_task(cron_scheduler_loop())
@app.get("/")
async def read_root():
    return FileResponse("static/index.html")

@app.get("/api/system/info")
def get_system_info():
    try:
        with open("version.txt", "r") as f:
            version = f.read().strip()
    except:
        version = "0.0.0"
        
    try:
        with open("CHANGELOG.md", "r") as f:
            changelog = f.read()
    except:
        changelog = "Changelog not found."
        
    return {"version": version, "changelog": changelog}

@app.get("/api/debug/db")
def debug_db():
    """Temporary debug endpoint to check database connection."""
    from database import DATABASE_URL
    db_url = DATABASE_URL
    result = {
        "database_url_set": bool(db_url),
        "database_url_preview": (db_url[:30] + "...") if db_url else None,
        "mode": "PostgreSQL" if db_url else "SQLite"
    }
    
    try:
        conn = get_db_connection()
        # Test query
        projects = conn.execute("SELECT id, name FROM projects").fetchall()
        result["projects_count"] = len(projects)
        result["projects"] = [dict(p) for p in projects]
        
        settings = conn.execute("SELECT key FROM settings").fetchall()
        result["settings_count"] = len(settings)
        result["settings_keys"] = [dict(s)['key'] for s in settings]
        
        conn.close()
        result["connection"] = "OK"
    except Exception as e:
        import traceback
        result["connection"] = "ERROR"
        result["error"] = str(e)
        result["traceback"] = traceback.format_exc()
    
    return result

@app.get("/api/settings")
def get_settings():
    conn = get_db_connection()
    rows = conn.execute("SELECT key, value FROM settings").fetchall()
    conn.close()
    return {row['key']: row['value'] for row in rows}

@app.post("/api/settings")
def update_settings(settings: GlobalSettings):
    conn = get_db_connection()
    try:
        if settings.service_account_json is not None:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                        ('service_account_json', settings.service_account_json))
        if settings.google_api_key is not None:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                        ('google_api_key', settings.google_api_key))
        if settings.default_sheet_id is not None:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                        ('default_sheet_id', settings.default_sheet_id))
        if settings.google_client_id is not None:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                        ('google_client_id', settings.google_client_id))
        if settings.google_client_secret is not None:
            conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", 
                        ('google_client_secret', settings.google_client_secret))
        conn.commit()
        return {"success": True, "message": "Settings updated"}
    except Exception as e:
        return {"success": False, "message": str(e)}
    finally:
        conn.close()

@app.post("/api/test_gemini")
async def test_gemini(request: dict = Body(...)):
    """
    Tests the Gemini API Key.
    """
    api_key = request.get('api_key')
    if not api_key:
        raise HTTPException(status_code=400, detail="Missing API Key")
        
    try:
        # 1. Update Env / Config temporarily for this test
        from google import genai
        client = genai.Client(api_key=api_key)
        
        # 2. Simple Generation
        response = client.models.generate_content(
            model="gemini-2.5-flash", 
            contents="Reply with 'OK' if you receive this."
        )
        
        if response.text:
            return {"success": True, "message": "Connection Successful!", "response": response.text}
        else:
             return {"success": False, "message": "No response text received."}
             
    except Exception as e:
        return {"success": False, "message": str(e)}

@app.websocket("/ws/workflow-status")
async def websocket_endpoint(websocket: WebSocket):
    await manager.connect(websocket)
    try:
        while True:
            await websocket.receive_text()
    except WebSocketDisconnect:
        manager.disconnect(websocket)

# OAuth Configuration
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive.readonly']
# Create a dummy client config if not provided, or load from DB.
# For now, we assume the user will provide Client ID/Secret in Global Settings.

def get_google_flow(state=None):
    conn = get_db_connection()
    settings = conn.execute("SELECT key, value FROM settings WHERE key IN ('google_client_id', 'google_client_secret')").fetchall()
    conn.close()
    
    settings_dict = {row['key']: row['value'] for row in settings}
    client_id = settings_dict.get('google_client_id')
    client_secret = settings_dict.get('google_client_secret')
    
    if not client_id or not client_secret:
        raise ValueError("Google Client ID and Secret not configured in Settings.")

    client_config = {
        "web": {
            "client_id": client_id,
            "client_secret": client_secret,
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
        }
    }
    
    flow = Flow.from_client_config(
        client_config,
        scopes=SCOPES,
        redirect_uri='http://localhost:8000/api/auth/google/callback'
    )
    
    if state:
        flow.state = state
        
    return flow

@app.get("/api/auth/google/url")
def get_google_auth_url(project_id: str):
    try:
        flow = get_google_flow()
        # Pass project_id in state so we know where to save tokens
        authorization_url, state = flow.authorization_url(
            access_type='offline',
            prompt='consent',
            include_granted_scopes='true',
            state=project_id 
        )
        return {"url": authorization_url}
    except Exception as e:
        return {"error": str(e)}

@app.get("/api/auth/google/callback")
def google_auth_callback(state: str, code: str):
    project_id = state
    try:
        flow = get_google_flow(state=state)
        flow.fetch_token(code=code)
        
        credentials = flow.credentials
        creds_json = {
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': credentials.scopes
        }
        
        conn = get_db_connection()
        conn.execute("UPDATE projects SET oauth_credentials = ? WHERE id = ?", 
                    (json.dumps(creds_json), project_id))
        conn.commit()
        conn.close()
        
        return HTMLResponse("<script>window.opener.postMessage('oauth_success', '*'); window.close();</script>")
    except Exception as e:
        return HTMLResponse(f"<h3>Error: {str(e)}</h3>")

@app.get("/api/projects/{project_id}/auth_status")
def get_project_auth_status(project_id: str):
    conn = get_db_connection()
    project = conn.execute("SELECT oauth_credentials FROM projects WHERE id = ?", (project_id,)).fetchone()
    conn.close()
    
    if not project or not project['oauth_credentials']:
        return {"connected": False}
    
    try:
        creds_data = json.loads(project['oauth_credentials'])
        creds = Credentials.from_authorized_user_info(creds_data, SCOPES)
        
        # Try to refresh if not valid
        if not creds.valid:
            if creds.refresh_token:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                # Save refreshed credentials
                conn = get_db_connection()
                conn.execute("UPDATE projects SET oauth_credentials = ? WHERE id = ?", (creds.to_json(), project_id))
                conn.commit()
                conn.close()
                return {"connected": True}
            else:
                return {"connected": False, "error": "Token expired, no refresh token. Please reconnect."}
        return {"connected": True}
    except Exception as e:
        print(f"Auth status check error: {e}")
        # Token is invalid/corrupted — clear it
        conn = get_db_connection()
        conn.execute("UPDATE projects SET oauth_credentials = NULL WHERE id = ?", (project_id,))
        conn.commit()
        conn.close()
        return {"connected": False, "error": f"Token invalid: {str(e)}. Please reconnect."}

@app.post("/api/projects/{project_id}/disconnect_google")
def disconnect_google(project_id: str):
    """Remove stored OAuth credentials for a project."""
    conn = get_db_connection()
    conn.execute("UPDATE projects SET oauth_credentials = NULL WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    return {"success": True, "message": "Google Account disconnected."}

@app.get("/api/projects/{project_id}/picker_token")
def get_picker_token(project_id: str):
    conn = get_db_connection()
    settings = conn.execute("SELECT value FROM settings WHERE key = 'google_client_id'").fetchone()
    client_id = settings['value'] if settings else None
    project = conn.execute("SELECT oauth_credentials FROM projects WHERE id = ?", (project_id,)).fetchone()
    conn.close()

    if not project or not project['oauth_credentials']:
        return {"token": None}

    try:
        creds_data = json.loads(project['oauth_credentials'])
        creds = Credentials.from_authorized_user_info(creds_data, SCOPES)
        
        # Always try to refresh if token is not valid
        if not creds.valid:
            if creds.refresh_token:
                from google.auth.transport.requests import Request
                creds.refresh(Request())
                # Save refreshed credentials back to DB
                conn = get_db_connection()
                conn.execute("UPDATE projects SET oauth_credentials = ? WHERE id = ?", (creds.to_json(), project_id))
                conn.commit()
                conn.close()
            else:
                return {"token": None, "error": "Token expired. Please reconnect Google Account."}
        
        return {"token": creds.token, "app_id": client_id}
    except Exception as e:
        print(f"Error getting picker token: {e}")
        return {"token": None, "error": str(e)}
import shutil
import os
from engine import process_price_list_file

@app.post("/api/projects/{project_id}/upload-price-list")
async def upload_price_list(project_id: str, file: UploadFile = File(...)):
    print(f"DEBUG: Received upload request for project {project_id}, file: {file.filename}")
    if not file.filename.endswith(('.xlsx', '.xls', '.csv')):
         return {"success": False, "message": "Only Excel (.xlsx, .xls) or CSV files are allowed."}
         
    # Use absolute path to avoid CWD issues
    base_dir = os.path.dirname(os.path.abspath(__file__))
    upload_dir = os.path.join(base_dir, "uploads", project_id)
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file.filename)
    
    try:
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        # Process and Cache
        result = process_price_list_file(project_id, file_path)
        return result
        
    except Exception as e:
        return {"success": False, "message": f"Upload failed: {str(e)}"}

@app.get("/api/projects")
def list_projects():
    conn = get_db_connection()
    projects = conn.execute("SELECT * FROM projects ORDER BY created_at DESC").fetchall()
    conn.close()
    return {"projects": [dict(p) for p in projects]}

@app.post("/api/projects")
def create_project(project: ProjectCreate):
    project_id = str(uuid.uuid4())
    conn = get_db_connection()
    try:
        conn.execute(
            "INSERT INTO projects (id, name, description, google_sheet_id, service_account_json, smtp_config, wesendit_config, pipedrive_config, cron_expression, price_list_url, locality_prompt, products_config, workflow_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            (
                project_id,
                project.name,
                project.description,
                project.google_sheet_id,
                project.service_account_json,
                json.dumps(project.smtp_config) if project.smtp_config else None,
                json.dumps(project.wesendit_config) if project.wesendit_config else None,
                json.dumps(project.pipedrive_config) if project.pipedrive_config else None,
                project.cron_expression,
                project.price_list_url,
                project.locality_prompt,
                json.dumps(project.products_config) if project.products_config else None,
                json.dumps(project.workflow_json) if project.workflow_json else None
            )
        )
        conn.commit()
        conn.commit()
    except Exception as e:
        conn.close()
        import traceback
        traceback.print_exc()
        print(f"ERROR Creating Project: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
    conn.close()
    return {"id": project_id, "message": "Project created successfully"}

@app.get("/api/projects/{project_id}")
def get_project(project_id: str):
    conn = get_db_connection()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    conn.close()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return dict(project)

@app.put("/api/projects/{project_id}")
def update_project(project_id: str, project: ProjectUpdate):
    conn = get_db_connection()
    current = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not current:
        conn.close()
        raise HTTPException(status_code=404, detail="Project not found")

    # Build update query dynamically
    fields = []
    values = []
    
    if project.name is not None:
        fields.append("name = ?")
        values.append(project.name)
    if project.description is not None:
        fields.append("description = ?")
        values.append(project.description)
    if project.status is not None:
        fields.append("status = ?")
        values.append(project.status)
    if project.google_sheet_id is not None:
        fields.append("google_sheet_id = ?")
        values.append(project.google_sheet_id)
    if project.service_account_json is not None:
        fields.append("service_account_json = ?")
        values.append(project.service_account_json)
    if project.cron_expression is not None:
        fields.append("cron_expression = ?")
        values.append(project.cron_expression)
    if project.price_list_url is not None:
        fields.append("price_list_url = ?")
        values.append(project.price_list_url)
    if project.locality_prompt is not None:
        fields.append("locality_prompt = ?")
        values.append(project.locality_prompt)
    if project.products_config is not None:
        fields.append("products_config = ?")
        values.append(json.dumps(project.products_config))
    if project.workflow_json is not None:
        fields.append("workflow_json = ?")
        values.append(json.dumps(project.workflow_json))
    if project.smtp_config is not None:
        fields.append("smtp_config = ?")
        values.append(json.dumps(project.smtp_config))
    if project.wesendit_config is not None:
        fields.append("wesendit_config = ?")
        values.append(json.dumps(project.wesendit_config))
    if project.pipedrive_config is not None:
        fields.append("pipedrive_config = ?")
        values.append(json.dumps(project.pipedrive_config))

    if not fields:
        conn.close()
        return {"message": "No changes to update"}

    values.append(project_id)
    query = f"UPDATE projects SET {', '.join(fields)} WHERE id = ?"
    
    try:
        conn.execute(query, tuple(values))
        
        # Auto-snapshot: save workflow version on each save if workflow_json changed
        if project.workflow_json is not None:
            wf_json_str = json.dumps(project.workflow_json)
            # Count existing versions to generate label
            count = conn.execute(
                "SELECT COUNT(*) FROM workflow_versions WHERE project_id = ?", (project_id,)
            ).fetchone()[0]
            version_label = f"v{count + 1}"
            conn.execute(
                "INSERT INTO workflow_versions (id, project_id, workflow_json, label) VALUES (?, ?, ?, ?)",
                (str(uuid.uuid4()), project_id, wf_json_str, version_label)
            )
            # Keep only last 20 versions (compatible with both SQLite and PostgreSQL)
            old_versions = conn.execute(
                "SELECT id FROM workflow_versions WHERE project_id = ? ORDER BY created_at DESC",
                (project_id,)
            ).fetchall()
            if len(old_versions) > 20:
                ids_to_delete = [v['id'] for v in old_versions[20:]]
                for vid in ids_to_delete:
                    conn.execute("DELETE FROM workflow_versions WHERE id = ?", (vid,))
        
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

    conn.close()
    return {"message": "Project updated successfully"}

# --- WORKFLOW VERSION HISTORY ---

@app.get("/api/projects/{project_id}/versions")
def list_versions(project_id: str):
    conn = get_db_connection()
    versions = conn.execute(
        "SELECT id, label, created_at FROM workflow_versions WHERE project_id = ? ORDER BY created_at DESC",
        (project_id,)
    ).fetchall()
    conn.close()
    return {"versions": [dict(v) for v in versions]}

@app.post("/api/projects/{project_id}/versions")
def create_snapshot(project_id: str, body: dict = Body({})):
    conn = get_db_connection()
    project = conn.execute("SELECT workflow_json FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project or not project['workflow_json']:
        conn.close()
        raise HTTPException(status_code=404, detail="No workflow to snapshot")
    
    count = conn.execute(
        "SELECT COUNT(*) FROM workflow_versions WHERE project_id = ?", (project_id,)
    ).fetchone()[0]
    label = body.get("label", f"Snapshot {count + 1}")
    
    version_id = str(uuid.uuid4())
    conn.execute(
        "INSERT INTO workflow_versions (id, project_id, workflow_json, label) VALUES (?, ?, ?, ?)",
        (version_id, project_id, project['workflow_json'], label)
    )
    conn.commit()
    conn.close()
    return {"success": True, "id": version_id, "label": label}

@app.post("/api/projects/{project_id}/versions/{version_id}/restore")
def restore_version(project_id: str, version_id: str):
    conn = get_db_connection()
    version = conn.execute(
        "SELECT workflow_json FROM workflow_versions WHERE id = ? AND project_id = ?",
        (version_id, project_id)
    ).fetchone()
    if not version:
        conn.close()
        raise HTTPException(status_code=404, detail="Version not found")
    
    conn.execute("UPDATE projects SET workflow_json = ? WHERE id = ?", (version['workflow_json'], project_id))
    conn.commit()
    conn.close()
    return {"success": True, "message": "Workflow restored"}

@app.delete("/api/projects/{project_id}")
def delete_project(project_id: str):
    conn = get_db_connection()
    conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))
    conn.commit()
    conn.close()
    return {"message": "Project deleted"}

# --- PROJECT DUPLICATE ---
@app.post("/api/projects/{project_id}/duplicate")
def duplicate_project(project_id: str):
    conn = get_db_connection()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    if not project:
        conn.close()
        raise HTTPException(status_code=404, detail="Project not found")
    
    p = dict(project)
    new_id = str(uuid.uuid4())
    new_name = (p.get('name') or 'Unnamed') + ' (Copy)'
    
    conn.execute("""
        INSERT INTO projects (id, name, description, google_sheet_id, service_account_json, 
            cron_expression, price_list_url, locality_prompt, products_config, workflow_json,
            smtp_config, wesendit_config, pipedrive_config, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        new_id, new_name, p.get('description', ''),
        p.get('google_sheet_id', ''), p.get('service_account_json', ''),
        p.get('cron_expression', ''), p.get('price_list_url', ''),
        p.get('locality_prompt', ''), p.get('products_config', ''),
        p.get('workflow_json', ''),
        p.get('smtp_config', ''), p.get('wesendit_config', ''),
        p.get('pipedrive_config', ''),
        'active'
    ))
    conn.commit()
    conn.close()
    return {"id": new_id, "name": new_name, "message": "Project duplicated"}

# --- KNOWLEDGE FILE PARSING ---
@app.post("/api/knowledge/parse")
async def parse_knowledge_file(file: UploadFile = File(...)):
    """Parse uploaded file (PDF, Excel, CSV, TXT, JSON, MD) and return extracted text."""
    try:
        content = await file.read()
        filename = file.filename.lower()
        text = ""
        
        if filename.endswith('.csv'):
            import csv
            import io
            decoded = content.decode('utf-8', errors='replace')
            reader = csv.reader(io.StringIO(decoded))
            rows = list(reader)
            text = '\n'.join([', '.join(row) for row in rows])
            
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
                import io
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
                all_text = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    all_text.append(f"--- Sheet: {sheet} ---")
                    for row in ws.iter_rows(values_only=True):
                        row_text = ', '.join([str(cell) if cell is not None else '' for cell in row])
                        if row_text.strip(', '):
                            all_text.append(row_text)
                text = '\n'.join(all_text)
                wb.close()
            except ImportError:
                text = "[ERROR] openpyxl not installed. Run: pip install openpyxl"
        
        elif filename.endswith('.pdf'):
            try:
                import pdfplumber
                import io
                with pdfplumber.open(io.BytesIO(content)) as pdf:
                    pages = []
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            pages.append(page_text)
                    text = '\n\n'.join(pages)
            except ImportError:
                try:
                    import PyPDF2
                    import io
                    reader = PyPDF2.PdfReader(io.BytesIO(content))
                    pages = []
                    for page in reader.pages:
                        page_text = page.extract_text()
                        if page_text:
                            pages.append(page_text)
                    text = '\n\n'.join(pages)
                except ImportError:
                    text = "[ERROR] pdfplumber or PyPDF2 not installed. Run: pip install pdfplumber"
        
        elif filename.endswith(('.txt', '.md', '.json')):
            text = content.decode('utf-8', errors='replace')
        
        else:
            return JSONResponse({"success": False, "error": f"Unsupported file type: {file.filename}"}, status_code=400)
        
        if not text.strip():
            return JSONResponse({"success": False, "error": "No text could be extracted from the file."}, status_code=400)
        
        return {"success": True, "text": text.strip(), "filename": file.filename, "chars": len(text.strip())}
        
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

from engine import test_project_connection, sync_price_list
from media_engine import save_uploaded_file, get_media_files, delete_media_file, ensure_project_dir

# --- MEDIA ROUTES ---
@app.get("/api/projects/{project_id}/media")
def list_media(project_id: str):
    return get_media_files(project_id)

@app.post("/api/projects/{project_id}/media/upload")
async def upload_media(project_id: str, file: UploadFile = File(...)):
    return await save_uploaded_file(project_id, file)

@app.delete("/api/projects/{project_id}/media/{filename}")
def delete_media(project_id: str, filename: str):
    return delete_media_file(project_id, filename)

@app.get("/api/projects/{project_id}/media/file/{filename}")
def get_media_file(project_id: str, filename: str):
    path = ensure_project_dir(project_id)
    file_path = os.path.join(path, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path)
    raise HTTPException(status_code=404, detail="File not found")

class SyncPriceRequest(BaseModel):
    sheet_id: str
    sheet_range: Optional[str] = "Foglio1!A:G"

@app.post("/api/projects/{project_id}/test")
def test_connection(project_id: str):
    result = test_project_connection(project_id)
    return result

@app.post("/api/projects/{project_id}/sync-prices")
def sync_prices(project_id: str, request: SyncPriceRequest):
    result = sync_price_list(project_id, request.sheet_id, request.sheet_range)
    if not result['success']:
        raise HTTPException(status_code=400, detail=result['message'])
    return result

@app.get("/api/projects/{project_id}/headers")
def get_project_headers(project_id: str):
    from engine import get_sheet_headers
    result = get_sheet_headers(project_id)
    if not result['success']:
         raise HTTPException(status_code=400, detail=result.get('message', 'Unknown error'))
    return result

# --- PRICE LIST UPLOAD PER PRODUCT ---
@app.post("/api/projects/{project_id}/product-price-list")
async def upload_product_price_list(project_id: str, product_index: int, file: UploadFile = File(...)):
    if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
         return JSONResponse({"success": False, "message": "Only Excel or CSV allowed"}, status_code=400)
    
    try:
        # Use media engine logic or custom path
        base_dir = os.path.dirname(os.path.abspath(__file__))
        upload_dir = os.path.join(base_dir, "uploads", project_id, "products")
        os.makedirs(upload_dir, exist_ok=True)
        
        filename = f"prod_{product_index}_{file.filename}"
        file_path = os.path.join(upload_dir, filename)
        
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
            
        return {"success": True, "filename": filename, "path": file_path}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# --- INTEGRATION TESTS ---
class EmailTestRequest(BaseModel):
    host: str
    port: int
    user: str
    password: str  # Changed from pass_ with alias
    from_name: str = "Floormad Manager"
    to_email: Optional[str] = None # Optional recipient

@app.post("/api/test/email")
def test_email_endpoint(req: EmailTestRequest):
    import requests as http_requests
    
    BRIDGE_URL = "http://workflow.floormad.com/bridge.php"
    
    print(f"DEBUG: Testing Email via Bridge with Host={req.host}, Port={req.port}, User={req.user}")
    
    # Try via bridge first (works on Railway)
    try:
        bridge_response = http_requests.post(BRIDGE_URL, json={
            "action": "send_test_email",
            "host": req.host,
            "port": req.port,
            "user": req.user,
            "password": req.password,
            "from_name": req.from_name,
            "to_email": req.to_email or req.user
        }, timeout=20)
        
        result = bridge_response.json()
        if result.get("success"):
            return {"success": True, "method": "bridge"}
        else:
            return JSONResponse({"success": False, "message": result.get("message", "Bridge error")}, status_code=500)
    
    except Exception as bridge_error:
        print(f"DEBUG: Bridge failed ({bridge_error}), trying direct SMTP...")
    
    # Fallback: direct SMTP (works locally)
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        msg = MIMEMultipart()
        msg['From'] = f"{req.from_name} <{req.user}>"
        recipient = req.to_email if req.to_email else req.user
        msg['To'] = recipient
        msg['Subject'] = "✅ Floormad - SMTP Test"
        msg.attach(MIMEText("SMTP connection test successful!", 'plain'))
        
        if int(req.port) == 465:
            server = smtplib.SMTP_SSL(req.host, req.port, timeout=10)
        else:
            server = smtplib.SMTP(req.host, req.port, timeout=10)
            server.starttls()
        
        server.login(req.user, req.password)
        server.send_message(msg)
        server.quit()
        return {"success": True, "method": "direct"}
    except Exception as e:
        print(f"DEBUG: Direct SMTP also failed: {e}")
        return JSONResponse({"success": False, "message": f"{type(e).__name__}: {str(e)}"}, status_code=500)


# WhatsApp Test Request Model
class WhatsAppTestRequest(BaseModel):
    api_key: str
    api_url: Optional[str] = None
    phone: str
    message: str = "Test message from Floormad"

@app.post("/api/test/whatsapp")
def test_whatsapp_endpoint(req: WhatsAppTestRequest):
    try:
        from wesender_client import WeSenderClient
        
        # Initialize Client
        client = WeSenderClient(req.api_key, req.api_url)
        
        # Send Message
        result = client.send_message(req.phone, req.message)
        
        if result.get('success'):
             return {"success": True, "api_response": result.get('data')}
        else:
             error_msg = result.get('details', {}).get('error', 'Unknown Error')
             return JSONResponse({"success": False, "message": f"WeSender Error: {error_msg}", "details": result.get('details')}, status_code=500)

    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# --- PIPEDRIVE TEST ---
class PipedriveTestRequest(BaseModel):
    api_token: str
    email: str = "test@example.com"

@app.post("/api/test/pipedrive")
def test_pipedrive_endpoint(req: PipedriveTestRequest):
    try:
        from pipedrive_client import PipedriveClient
        
        client = PipedriveClient(req.api_token)
        person = client.search_person(req.email)
        
        if person:
            return {"success": True, "message": f"Connected! Found: {person.get('name', 'Unknown')}", "person": person}
        else:
            return {"success": True, "message": "Connected to Pipedrive! No person found with that email (this is OK)."}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# Duplicate read_root removed — defined at the top of the file
@app.post("/api/projects/{project_id}/run")
async def run_project_workflow(project_id: str):
    conn = get_db_connection()
    project = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
    
    # Global Settings for API Keys
    settings = conn.execute("SELECT key, value FROM settings").fetchall()
    settings_dict = {row['key']: row['value'] for row in settings}
    conn.close()
    
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    workflow_json = project['workflow_json']
    if not workflow_json:
         return {"success": False, "message": "No workflow defined"}

    try:
        workflow_data = json.loads(workflow_json)
        
        # Callback for real-time updates
        async def status_callback(node_id, status, message=""):
            await manager.broadcast(json.dumps({
                "type": "execution_update",
                "node_id": node_id,
                "status": status,
                "message": message
            }))

        engine = WorkflowEngine(
            workflow_data, 
            context={"project": dict(project)},
            api_key=settings_dict.get('google_api_key')
        )
        
        # Run synchronous engine with async callback wrapper? 
        # engine.run is sync. We need to make it async or use a sync wrapper that calls async?
        # Simpler: Make engine.run accept a sync callback, and we define a sync wrapper that uses run_coroutine_threadsafe or just simple printing if standard WS is async.
        # Actually, FastAPI handlers are async. 
        # Let's pass the manager and let engine run. But engine.run is sync.
        # We will modify engine.run to be async for better IO handling! 
        # OR keeping it sync, we can't await `manager.broadcast`.
        # Correct path: Update WorkflowEngine.run to be `async def run(...)`.
        
        result = await engine.run(status_callback) 
        
        # Sanitize result to prevent circular references in JSON serialization
        # This is critical: accumulated_context can contain nested references
        try:
            sanitized = json.loads(json.dumps(result, default=str))
            return sanitized
        except Exception as json_err:
            logger.error(f"JSON sanitization failed: {json_err}")
            # Ultimate fallback: return only safe parts
            safe_result = {
                "status": result.get("status", "unknown"),
                "log": result.get("log", []),
                "final_context": {}
            }
            # Try to salvage individual context keys
            ctx = result.get("final_context", {})
            for k, v in ctx.items():
                try:
                    json.dumps(v, default=str)
                    safe_result["final_context"][k] = v
                except:
                    safe_result["final_context"][k] = str(v)[:500]
            return safe_result
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"Workflow execution error:\n{tb}")
        return {
            "status": "failed",
            "success": False,
            "message": str(e),
            "traceback": tb,
            "log": engine.execution_log if 'engine' in dir() else []
        }

@app.post("/api/projects/{project_id}/optimize_file")
def run_file_optimization(project_id: str, request: dict = Body(...)):
    """
    Optimizes a confirmed file (Price List or Knowledge Base).
    Body: { "file_name": "...", "type": "price_list" | "knowledge_base" }
    """
    file_name = request.get('file_name')
    file_type = request.get('type')
    
    if not file_name or not file_type:
        raise HTTPException(status_code=400, detail="Missing file_name or type")
        
    from engine import optimize_price_list_with_ai, optimize_knowledge_base_with_ai
    
    if file_type == 'price_list':
        result = optimize_price_list_with_ai(project_id, file_name)
    elif file_type == 'knowledge_base':
        result = optimize_knowledge_base_with_ai(project_id, file_name)
    else:
         return {"success": False, "message": "Unknown type"}
         
    return result

@app.get("/api/projects/{project_id}/read_file")
def read_project_file(project_id: str, file: str):
    """
    Reads a file from the project's upload directory.
    """
    if not file or ".." in file or "/" in file: # Basic security check
         raise HTTPException(status_code=400, detail="Invalid filename")

    import os
    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    possible_paths = [
        os.path.join(base_dir, "uploads", str(project_id), file),
        os.path.join(os.getcwd(), "uploads", str(project_id), file),
        os.path.join(base_dir, "..", "uploads", str(project_id), file)
    ]
    
    file_path = None
    for p in possible_paths:
        if os.path.exists(p):
            file_path = p
            break
            
    if not file_path:
        raise HTTPException(status_code=404, detail=f"File not found: {file}")
        
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            return f.read()
    except Exception as e:
         raise HTTPException(status_code=500, detail=f"Error reading file: {e}")

# --- DATABASE BACKUP & RESTORE ---
from datetime import datetime as dt_backup

@app.get("/api/backup")
def download_backup():
    """Download a full JSON backup of the database."""
    conn = get_db_connection()
    try:
        projects = conn.execute("SELECT * FROM projects").fetchall()
        settings_rows = conn.execute("SELECT * FROM settings").fetchall()
        runs = conn.execute("SELECT * FROM runs").fetchall()
        try:
            versions = conn.execute("SELECT * FROM workflow_versions").fetchall()
        except:
            versions = []
        conn.close()
        
        backup = {
            "backup_date": dt_backup.now().isoformat(),
            "projects": [dict(p) for p in projects],
            "settings": [dict(s) for s in settings_rows],
            "runs": [dict(r) for r in runs],
            "workflow_versions": [dict(v) for v in versions]
        }
        
        return JSONResponse(content=backup, headers={
            "Content-Disposition": f"attachment; filename=floormad_backup_{dt_backup.now().strftime('%Y%m%d_%H%M%S')}.json"
        })
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/backup/restore")
async def restore_backup(file: UploadFile = File(...)):
    """Restore database from a JSON backup file."""
    try:
        content = await file.read()
        backup = json.loads(content)
        conn = get_db_connection()
        restored = {"projects": 0, "settings": 0, "runs": 0, "versions": 0}
        
        for s in backup.get("settings", []):
            try:
                conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (s['key'], s.get('value')))
                restored["settings"] += 1
            except: pass
        
        for p in backup.get("projects", []):
            cols = ['id', 'name', 'description', 'status', 'google_sheet_id', 'service_account_json',
                    'smtp_config', 'wesendit_config', 'cron_expression', 'price_list_url',
                    'locality_prompt', 'products_config', 'workflow_json', 'oauth_credentials',
                    'pipedrive_config', 'price_list_cache']
            available_cols = [c for c in cols if c in p]
            vals = [p.get(c, '' if c in ('google_sheet_id', 'service_account_json') else None) for c in available_cols]
            placeholders = ', '.join(['?' for _ in available_cols])
            try:
                conn.execute(f"INSERT INTO projects ({', '.join(available_cols)}) VALUES ({placeholders})", tuple(vals))
                restored["projects"] += 1
            except: pass
        
        for r in backup.get("runs", []):
            try:
                conn.execute("INSERT INTO runs (id, project_id, timestamp, leads_processed, status, log_details) VALUES (?, ?, ?, ?, ?, ?)",
                    (r['id'], r.get('project_id'), r.get('timestamp'), r.get('leads_processed', 0), r.get('status'), r.get('log_details')))
                restored["runs"] += 1
            except: pass
        
        for v in backup.get("workflow_versions", []):
            try:
                conn.execute("INSERT INTO workflow_versions (id, project_id, workflow_json, label, created_at) VALUES (?, ?, ?, ?, ?)",
                    (v['id'], v['project_id'], v['workflow_json'], v.get('label'), v.get('created_at')))
                restored["versions"] += 1
            except: pass
        
        conn.commit()
        conn.close()
        return {"success": True, "message": "Backup restored!", "restored": restored}
    except Exception as e:
        import traceback
        return {"success": False, "message": str(e), "traceback": traceback.format_exc()}

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    print(f"Starting server on 0.0.0.0:{port}")
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=False)
